from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func

from ..auth import get_current_user
from ..database import get_db
from ..config import settings
from ..models import User, Restaurant, Order, MenuItem, OperatorMember, RestaurantReview, RestaurantImage, MenuCategory, ModifierGroup, ModifierOption, MenuItemModifierGroup, RestaurantStation
from ..schemas import (
    RestaurantOut,
    OrdersListOut,
    OrderOut,
    OrderItemOut,
    MenuItemOut,
    RestaurantImageCreateIn,
    RestaurantImageOut,
    RestaurantHoursIn,
    CategoryIn,
    CategoryOut,
    ModifierGroupIn,
    ModifierGroupOut,
    ModifierOptionIn,
    ModifierOptionOut,
)
from ..utils.notify import notify
from ..utils.webhooks import send_webhooks
from ..utils.webhooks import process_pending_once
from ..utils.audit import audit
from ..utils.print_escpos import print_ticket


router = APIRouter(prefix="/operator", tags=["operator"])


def _ensure_operator(db: Session, user_id: str, min_role: str = "agent") -> OperatorMember:
    mem = db.query(OperatorMember).filter(OperatorMember.user_id == user_id).one_or_none()
    if mem is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not operator member")
    # Role ranking: admin > manager > agent > kitchen > cashier > support
    ranks = {"support": 0, "cashier": 1, "kitchen": 2, "agent": 3, "manager": 4, "admin": 5}
    need = ranks.get(min_role, 0)
    have = ranks.get(mem.role or "", 0)
    if have < need:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=f"Role {min_role} required")
    return mem


@router.post("/dev/become_admin")
def dev_become_admin(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if settings.ENV != "dev":
        raise HTTPException(status_code=403, detail="Disabled outside dev")
    mem = db.query(OperatorMember).filter(OperatorMember.user_id == user.id).one_or_none()
    if mem is None:
        db.add(OperatorMember(user_id=user.id, role="admin"))
        db.flush()
    return {"detail": "ok"}


@router.get("/restaurants", response_model=list[RestaurantOut])
def list_restaurants(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = db.query(Restaurant).order_by(Restaurant.created_at.desc()).limit(500).all()
    out: list[RestaurantOut] = []
    for r in rows:
        avg, cnt = (
            db.query(func.avg(RestaurantReview.rating), func.count(RestaurantReview.id))
            .filter(RestaurantReview.restaurant_id == r.id)
            .one()
        )
        hours_dict = None
        is_open = None
        if r.hours_json:
            try:
                import json as _json
                hours_dict = _json.loads(r.hours_json)
            except Exception:
                hours_dict = None
        is_open = _compute_is_open(hours_dict, r.is_open_override)
        out.append(
            RestaurantOut(
                id=str(r.id),
                name=r.name,
                city=r.city,
                description=r.description,
                address=r.address,
                rating_avg=float(avg) if avg is not None else None,
                rating_count=int(cnt) if cnt is not None else 0,
                is_open=is_open,
                hours=hours_dict,
            )
        )
    return out


@router.post("/restaurants")
def create_restaurant(name: str, city: str | None = None, address: str | None = None, owner_phone: str | None = None, hours_json: str | None = None, is_open_override: bool | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    r = Restaurant(name=name, city=city, address=address, hours_json=hours_json, is_open_override=is_open_override)
    if owner_phone:
        owner = db.query(User).filter(User.phone == owner_phone).one_or_none()
        if owner is None:
            owner = User(phone=owner_phone)
            db.add(owner)
            db.flush()
        r.owner_user_id = owner.id
    db.add(r)
    db.flush()
    return {"id": str(r.id)}


@router.patch("/restaurants/{restaurant_id}")
def update_restaurant(restaurant_id: str, name: str | None = None, city: str | None = None, description: str | None = None, address: str | None = None, owner_phone: str | None = None, hours_json: str | None = None, is_open_override: bool | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    r = db.get(Restaurant, restaurant_id)
    if not r:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    if name is not None:
        r.name = name
    if city is not None:
        r.city = city
    if description is not None:
        r.description = description
    if address is not None:
        r.address = address
    if hours_json is not None:
        r.hours_json = hours_json
    if is_open_override is not None:
        r.is_open_override = is_open_override
    if owner_phone is not None:
        owner = db.query(User).filter(User.phone == owner_phone).one_or_none()
        if owner is None:
            owner = User(phone=owner_phone)
            db.add(owner)
            db.flush()
        r.owner_user_id = owner.id
    return {"detail": "ok"}


@router.get("/orders", response_model=OrdersListOut)
def list_all_orders(status_filter: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    q = db.query(Order).order_by(Order.created_at.desc())
    if status_filter:
        q = q.filter(Order.status == status_filter)
    rows = q.limit(500).all()
    return OrdersListOut(orders=[
        OrderOut(
            id=str(o.id), status=o.status, restaurant_id=str(o.restaurant_id), total_cents=o.total_cents,
            delivery_address=o.delivery_address, created_at=o.created_at, payment_request_id=o.payment_request_id,
            items=[OrderItemOut(menu_item_id=str(oi.menu_item_id), name=oi.name_snapshot, qty=oi.qty, price_cents=oi.price_cents_snapshot, subtotal_cents=oi.subtotal_cents) for oi in o.items]
        ) for o in rows
    ])


@router.get("/reports/summary")
def summary(days: int = 7, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow()
    fr = to - timedelta(days=max(1, min(days, 90)))
    total_orders = db.query(func.count(Order.id)).filter(Order.created_at >= fr, Order.created_at <= to).scalar() or 0
    delivered = db.query(func.count(Order.id)).filter(Order.status == "delivered", Order.created_at >= fr, Order.created_at <= to).scalar() or 0
    revenue = db.query(func.sum(Order.total_cents)).filter(Order.status == "delivered", Order.created_at >= fr, Order.created_at <= to).scalar() or 0
    return {
        "from_utc": fr.isoformat() + "Z",
        "to_utc": to.isoformat() + "Z",
        "orders_total": int(total_orders),
        "orders_delivered": int(delivered),
        "gross_revenue_cents": int(revenue or 0),
    }

from fastapi import Response
import csv
from io import StringIO
from io import BytesIO
import datetime as _dt
try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None
from fastapi import Request


@router.get("/reports/summary.csv")
def summary_csv(days: int = 7, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    data = summary(days, user, db)
    buf = StringIO(); w = csv.writer(buf)
    w.writerow(["from_utc","to_utc","orders_total","orders_delivered","gross_revenue_cents"])
    w.writerow([data["from_utc"], data["to_utc"], data["orders_total"], data["orders_delivered"], data["gross_revenue_cents"]])
    return Response(buf.getvalue(), media_type="text/csv")


@router.get("/reports/summary.pdf")
def summary_pdf(days: int = 7, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    data = summary(days, user, db)
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
    except Exception as e:
        # Dependency not available
        raise HTTPException(status_code=500, detail=f"PDF generator unavailable: {e}")
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, y, "Food Operator Report — Summary")
    c.setFont("Helvetica", 12)
    y -= 24
    c.drawString(72, y, f"Range: {data['from_utc']} .. {data['to_utc']}")
    y -= 18
    c.drawString(72, y, f"Orders Total: {data['orders_total']}")
    y -= 18
    c.drawString(72, y, f"Orders Delivered: {data['orders_delivered']}")
    y -= 18
    c.drawString(72, y, f"Gross Revenue (cents): {data['gross_revenue_cents']}")
    c.showPage(); c.save()
    pdf = buf.getvalue()
    return Response(pdf, media_type="application/pdf")


@router.get("/reports/summary.xlsx")
def summary_xlsx(days: int = 7, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX export unavailable (openpyxl not installed)")
    data = summary(days, user, db)
    from openpyxl import Workbook  # type: ignore
    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    ws.append(["from_utc","to_utc","orders_total","orders_delivered","gross_revenue_cents"])
    ws.append([data["from_utc"], data["to_utc"], data["orders_total"], data["orders_delivered"], data["gross_revenue_cents"]])
    out = BytesIO(); wb.save(out)
    return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# --- Hours helpers & endpoints ---


def _compute_is_open(hours: dict | None, override: bool | None, specials: dict | None = None) -> bool | None:
    if override is not None:
        return bool(override)
    if not hours:
        return None
    try:
        import datetime as _dt
        now = _dt.datetime.utcnow()
        if specials:
            key = now.strftime("%Y-%m-%d")
            if key in specials:
                slots = specials.get(key) or []
                if slots in ([], None, "closed"):
                    return False
                hm = now.hour * 60 + now.minute
                for s in slots:
                    if not isinstance(s, (list, tuple)) or len(s) != 2:
                        continue
                    a, b = s
                    ah, am = map(int, str(a).split(":"))
                    bh, bm = map(int, str(b).split(":"))
                    if (ah*60+am) <= hm <= (bh*60+bm):
                        return True
                return False
        weekday = ["mon","tue","wed","thu","fri","sat","sun"][now.weekday()]
        slots = hours.get(weekday) or []
        hm = now.hour * 60 + now.minute
        for s in slots:
            if not isinstance(s, (list, tuple)) or len(s) != 2:
                continue
            a, b = s
            ah, am = map(int, str(a).split(":"))
            bh, bm = map(int, str(b).split(":"))
            if (ah*60+am) <= hm <= (bh*60+bm):
                return True
        return False
    except Exception:
        return None


@router.get("/restaurants/{restaurant_id}/hours")
def get_hours(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    r = db.get(Restaurant, restaurant_id)
    if not r:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Restaurant not found")
    import json as _json
    hours = None
    if r.hours_json:
        try:
            hours = _json.loads(r.hours_json)
        except Exception:
            hours = None
    specials = None
    if getattr(r, "special_hours_json", None):
        try:
            specials = _json.loads(r.special_hours_json)
        except Exception:
            specials = None
    return {"hours": hours, "special_hours": specials, "is_open_override": r.is_open_override, "is_open": _compute_is_open(hours, r.is_open_override, specials)}


@router.post("/restaurants/{restaurant_id}/hours")
def set_hours(restaurant_id: str, payload: RestaurantHoursIn, is_open_override: bool | None = None, special_hours_json: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    r = db.get(Restaurant, restaurant_id)
    if not r:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Restaurant not found")
    import json as _json
    r.hours_json = _json.dumps(payload.hours)
    if special_hours_json is not None:
        r.special_hours_json = special_hours_json
    if is_open_override is not None:
        r.is_open_override = is_open_override
    return {"detail": "ok"}


# --- New: Restaurant menu management (platform operator) ---


@router.get("/restaurants/{restaurant_id}/menu", response_model=list[MenuItemOut])
def op_list_menu(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = (
        db.query(MenuItem)
        .filter(MenuItem.restaurant_id == restaurant_id)
        .order_by(MenuItem.created_at.desc())
        .all()
    )
    return [
        MenuItemOut(
            id=str(m.id),
            restaurant_id=str(m.restaurant_id),
            name=m.name,
            description=m.description,
            price_cents=m.price_cents,
            available=m.available,
            visible=getattr(m, 'visible', None),
            category_id=str(getattr(m, 'category_id', '') or '') or None,
            stock_qty=getattr(m, 'stock_qty', None),
            oos_until=getattr(m, 'oos_until', None),
            station=getattr(m, 'station', None),
        )
        for m in rows
    ]


@router.get("/restaurants/{restaurant_id}/items.csv")
def export_items_csv(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).order_by(MenuItem.created_at.desc()).all()
    import csv
    from io import StringIO
    buf = StringIO(); w = csv.writer(buf)
    w.writerow(["id","name","price_cents","available","visible","stock_qty","oos_until","station"])
    for m in rows:
        w.writerow([
            str(m.id),
            m.name or "",
            int(m.price_cents or 0),
            "true" if m.available else "false",
            "true" if getattr(m, 'visible', True) else "false",
            "" if getattr(m, 'stock_qty', None) is None else int(getattr(m, 'stock_qty')),
            "" if getattr(m, 'oos_until', None) is None else getattr(m, 'oos_until').isoformat()+"Z",
            getattr(m, 'station', None) or "",
        ])
    return Response(buf.getvalue(), media_type="text/csv")


@router.get("/restaurants/{restaurant_id}/items.xlsx")
def export_items_xlsx(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX export unavailable")
    rows = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).order_by(MenuItem.created_at.desc()).all()
    from openpyxl import Workbook  # type: ignore
    wb = Workbook(); ws = wb.active; ws.title = "Items"
    ws.append(["id","name","price_cents","available","visible","stock_qty","oos_until","station"])
    for m in rows:
        ws.append([
            str(m.id),
            m.name or "",
            int(m.price_cents or 0),
            bool(m.available),
            bool(getattr(m, 'visible', True)),
            "" if getattr(m, 'stock_qty', None) is None else int(getattr(m, 'stock_qty')),
            "" if getattr(m, 'oos_until', None) is None else getattr(m, 'oos_until').isoformat()+"Z",
            getattr(m, 'station', None) or "",
        ])
    out = BytesIO(); wb.save(out)
    return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/restaurants/{restaurant_id}/menu")
def op_create_menu_item(
    restaurant_id: str,
    name: str,
    price_cents: int,
    description: str | None = None,
    available: bool = True,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_operator(db, user.id, min_role="manager")
    r = db.get(Restaurant, restaurant_id)
    if not r:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Restaurant not found")
    mi = MenuItem(
        restaurant_id=r.id,
        name=name,
        description=description,
        price_cents=price_cents,
        available=available,
    )
    db.add(mi)
    db.flush()
    return {"id": str(mi.id)}


@router.patch("/menu/{menu_item_id}")
def op_update_menu_item(
    menu_item_id: str,
    name: str | None = None,
    price_cents: int | None = None,
    description: str | None = None,
    available: bool | None = None,
    visible: bool | None = None,
    category_id: str | None = None,
    stock_qty: int | None = None,
    oos_until: str | None = None,
    station: str | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_operator(db, user.id, min_role="manager")
    mi = db.get(MenuItem, menu_item_id)
    if not mi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Menu item not found")
    before = {"name": mi.name, "price_cents": mi.price_cents, "available": mi.available, "visible": getattr(mi, 'visible', None), "category_id": str(getattr(mi, 'category_id', '') or ''), "stock_qty": getattr(mi, 'stock_qty', None), "oos_until": str(getattr(mi, 'oos_until', None))}
    if name is not None:
        mi.name = name
    if price_cents is not None:
        mi.price_cents = price_cents
    if description is not None:
        mi.description = description
    if available is not None:
        mi.available = available
    if visible is not None:
        mi.visible = visible
    if category_id is not None:
        mi.category_id = category_id  # no strict FK validation here
    if stock_qty is not None:
        mi.stock_qty = stock_qty
    if oos_until is not None:
        try:
            from datetime import datetime as _dt
            mi.oos_until = _dt.fromisoformat(oos_until.replace("Z", "+00:00"))
        except Exception:
            pass
    if station is not None:
        if station == "":
            mi.station = None
        else:
            ok = db.query(RestaurantStation).filter(RestaurantStation.restaurant_id == mi.restaurant_id, RestaurantStation.name == station).one_or_none()
            if not ok:
                raise HTTPException(status_code=400, detail="Invalid station")
            mi.station = station
    audit(db, user_id=str(user.id), action="menu_item.update", entity_type="menu_item", entity_id=str(mi.id), before=before, after={"name": mi.name, "price_cents": mi.price_cents, "available": mi.available, "visible": getattr(mi, 'visible', None), "category_id": str(getattr(mi, 'category_id', '') or ''), "stock_qty": getattr(mi, 'stock_qty', None), "oos_until": str(getattr(mi, 'oos_until', None)), "station": getattr(mi, 'station', None)})
    return {"detail": "ok"}


@router.delete("/menu/{menu_item_id}")
def op_delete_menu_item(menu_item_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    mi = db.get(MenuItem, menu_item_id)
    if not mi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Menu item not found")
    before = {"name": mi.name}
    db.delete(mi)
    audit(db, user_id=str(user.id), action="menu_item.delete", entity_type="menu_item", entity_id=str(menu_item_id), before=before, after=None)
    return {"detail": "ok"}


# --- New: Restaurant images management (platform operator) ---


@router.post("/restaurants/{restaurant_id}/images", response_model=list[RestaurantImageOut])
def op_add_images(
    restaurant_id: str,
    images: list[RestaurantImageCreateIn],
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_operator(db, user.id, min_role="manager")
    r = db.get(Restaurant, restaurant_id)
    if not r:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Restaurant not found")
    created: list[RestaurantImageOut] = []
    for im in images:
        pi = RestaurantImage(restaurant_id=r.id, url=im.url, sort_order=im.sort_order)
        db.add(pi)
        db.flush()
        created.append(RestaurantImageOut(id=str(pi.id), url=pi.url, sort_order=pi.sort_order))
    audit(db, user_id=str(user.id), action="restaurant.images.create", entity_type="restaurant", entity_id=str(restaurant_id), before=None, after=[i.model_dump() for i in created])
    return created


@router.get("/restaurants/{restaurant_id}/images", response_model=list[RestaurantImageOut])
def op_list_images(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    imgs = (
        db.query(RestaurantImage)
        .filter(RestaurantImage.restaurant_id == restaurant_id)
        .order_by(RestaurantImage.sort_order.asc(), RestaurantImage.created_at.asc())
        .all()
    )
    return [RestaurantImageOut(id=str(i.id), url=i.url, sort_order=i.sort_order) for i in imgs]


@router.delete("/images/{image_id}")
def op_delete_image(image_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    img = db.get(RestaurantImage, image_id)
    if not img:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not found")
    db.delete(img)
    audit(db, user_id=str(user.id), action="restaurant.images.delete", entity_type="restaurant_image", entity_id=str(image_id), before={"url": img.url}, after=None)
    return {"detail": "ok"}


# --- Stations management ---


@router.get("/restaurants/{restaurant_id}/stations")
def list_stations(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = db.query(RestaurantStation).filter(RestaurantStation.restaurant_id == restaurant_id).order_by(RestaurantStation.sort_order.asc(), RestaurantStation.created_at.asc()).all()
    return [{"id": str(s.id), "name": s.name, "sort_order": s.sort_order} for s in rows]


@router.post("/restaurants/{restaurant_id}/stations")
def create_station(restaurant_id: str, name: str, sort_order: int = 0, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    s = RestaurantStation(restaurant_id=restaurant_id, name=name, sort_order=sort_order)
    db.add(s); db.flush()
    audit(db, user_id=str(user.id), action="station.create", entity_type="station", entity_id=str(s.id), before=None, after={"name": name, "sort_order": sort_order})
    return {"id": str(s.id)}


@router.delete("/stations/{station_id}")
def delete_station(station_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    s = db.get(RestaurantStation, station_id)
    if not s:
        raise HTTPException(status_code=404, detail="Not found")
    audit(db, user_id=str(user.id), action="station.delete", entity_type="station", entity_id=str(s.id), before={"name": s.name}, after=None)
    db.delete(s)
    return {"detail": "ok"}


# --- Categories ---


@router.get("/restaurants/{restaurant_id}/categories", response_model=list[CategoryOut])
def list_categories(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = db.query(MenuCategory).filter(MenuCategory.restaurant_id == restaurant_id).order_by(MenuCategory.sort_order.asc(), MenuCategory.created_at.asc()).all()
    return [CategoryOut(id=str(c.id), name=c.name, parent_id=str(c.parent_id) if c.parent_id else None, description=c.description, sort_order=c.sort_order) for c in rows]


@router.post("/restaurants/{restaurant_id}/categories")
def create_category(restaurant_id: str, payload: CategoryIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    c = MenuCategory(restaurant_id=restaurant_id, name=payload.name, parent_id=payload.parent_id, description=payload.description, sort_order=payload.sort_order)
    db.add(c); db.flush()
    audit(db, user_id=str(user.id), action="category.create", entity_type="category", entity_id=str(c.id), before=None, after=payload.model_dump())
    return {"id": str(c.id)}


@router.patch("/categories/{category_id}")
def update_category(category_id: str, payload: CategoryIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    c = db.get(MenuCategory, category_id)
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    before = {"name": c.name, "parent_id": str(c.parent_id) if c.parent_id else None, "description": c.description, "sort_order": c.sort_order}
    c.name = payload.name; c.parent_id = payload.parent_id; c.description = payload.description; c.sort_order = payload.sort_order
    audit(db, user_id=str(user.id), action="category.update", entity_type="category", entity_id=str(c.id), before=before, after=payload.model_dump())
    return {"detail": "ok"}


@router.delete("/categories/{category_id}")
def delete_category(category_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    c = db.get(MenuCategory, category_id)
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    audit(db, user_id=str(user.id), action="category.delete", entity_type="category", entity_id=str(c.id), before={"name": c.name}, after=None)
    db.delete(c)
    return {"detail": "ok"}


# --- Modifiers ---


@router.get("/restaurants/{restaurant_id}/modifier_groups", response_model=list[ModifierGroupOut])
def list_modifier_groups(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    rows = db.query(ModifierGroup).filter(ModifierGroup.restaurant_id == restaurant_id).order_by(ModifierGroup.sort_order.asc(), ModifierGroup.created_at.asc()).all()
    return [ModifierGroupOut(id=str(g.id), name=g.name, min_choices=g.min_choices, max_choices=g.max_choices, required=g.required, sort_order=g.sort_order) for g in rows]


@router.post("/restaurants/{restaurant_id}/modifier_groups")
def create_modifier_group(restaurant_id: str, payload: ModifierGroupIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    g = ModifierGroup(restaurant_id=restaurant_id, name=payload.name, min_choices=payload.min_choices, max_choices=payload.max_choices, required=payload.required, sort_order=payload.sort_order)
    db.add(g); db.flush()
    audit(db, user_id=str(user.id), action="modifier_group.create", entity_type="modifier_group", entity_id=str(g.id), before=None, after=payload.model_dump())
    return {"id": str(g.id)}


@router.post("/modifier_groups/{group_id}/options")
def create_modifier_option(group_id: str, payload: ModifierOptionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    g = db.get(ModifierGroup, group_id)
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")
    o = ModifierOption(group_id=g.id, name=payload.name, price_delta_cents=payload.price_delta_cents, sort_order=payload.sort_order)
    db.add(o); db.flush()
    audit(db, user_id=str(user.id), action="modifier_option.create", entity_type="modifier_option", entity_id=str(o.id), before=None, after=payload.model_dump())
    return {"id": str(o.id)}


@router.post("/menu/{menu_item_id}/modifier_groups/{group_id}")
def attach_modifier_group(menu_item_id: str, group_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    mi = db.get(MenuItem, menu_item_id)
    g = db.get(ModifierGroup, group_id)
    if not mi or not g:
        raise HTTPException(status_code=404, detail="Not found")
    link = db.query(MenuItemModifierGroup).filter(MenuItemModifierGroup.menu_item_id == mi.id, MenuItemModifierGroup.group_id == g.id).one_or_none()
    if not link:
        db.add(MenuItemModifierGroup(menu_item_id=mi.id, group_id=g.id))
        audit(db, user_id=str(user.id), action="menu_item.attach_group", entity_type="menu_item", entity_id=str(mi.id), before=None, after={"group_id": str(g.id)})
    return {"detail": "ok"}


@router.delete("/menu/{menu_item_id}/modifier_groups/{group_id}")
def detach_modifier_group(menu_item_id: str, group_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    link = db.query(MenuItemModifierGroup).filter(MenuItemModifierGroup.menu_item_id == menu_item_id, MenuItemModifierGroup.group_id == group_id).one_or_none()
    if link:
        db.delete(link)
        audit(db, user_id=str(user.id), action="menu_item.detach_group", entity_type="menu_item", entity_id=str(menu_item_id), before={"group_id": str(group_id)}, after=None)
    return {"detail": "ok"}


# --- Bulk orders ---


@router.post("/orders/bulk_status")
def bulk_status(order_ids: list[str], status_value: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="manager")
    changed = 0
    for oid in order_ids:
        o = db.get(Order, oid)
        if not o or o.status in ("delivered", "canceled"):
            continue
        allowed = ALLOWED_TRANSITIONS.get(o.status, set())
        if status_value not in allowed:
            continue
        prev = o.status
        o.status = status_value
        changed += 1
        audit(db, user_id=str(user.id), action="order.bulk_status", entity_type="order", entity_id=str(oid), before={"status": prev}, after={"status": status_value})
    return {"updated": changed}


# --- Bulk stock/OOS import (CSV) ---


@router.post("/restaurants/{restaurant_id}/bulk_stock")
def bulk_stock(restaurant_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db), request: __import__('fastapi').Request = None, format: str = 'json'):
    _ensure_operator(db, user.id, min_role="manager")
    body = (request and __import__('asyncio').get_event_loop().run_until_complete(request.body())) if request else b""
    if not body:
        raise HTTPException(status_code=400, detail="Empty body")
    txt = body.decode('utf-8', 'ignore')
    import csv, io
    reader = csv.DictReader(io.StringIO(txt))
    updated = 0
    details: list[dict] = []
    for idx, row in enumerate(reader, start=2):
        mid = (row.get('id') or '').strip()
        name = (row.get('name') or '').strip()
        mi = None
        if mid:
            mi = db.get(MenuItem, mid)
        if not mi and name:
            mi = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id, MenuItem.name == name).one_or_none()
        if not mi:
            details.append({"row": row, "index": idx, "status": "skipped", "reason": "item_not_found"})
            continue
        try:
            if 'stock_qty' in row and row['stock_qty'] != '':
                mi.stock_qty = int(row['stock_qty'])
            if 'available' in row and row['available'] != '':
                mi.available = str(row['available']).lower() in ('1','true','yes','y')
            if 'visible' in row and row['visible'] != '':
                mi.visible = str(row['visible']).lower() in ('1','true','yes','y')
            if 'oos_until' in row and row['oos_until']:
                try:
                    from datetime import datetime as _dt
                    mi.oos_until = _dt.fromisoformat(row['oos_until'].replace('Z','+00:00'))
                except Exception:
                    pass
            if 'station' in row and row['station']:
                st_name = row['station'].strip()
                ok = db.query(RestaurantStation).filter(RestaurantStation.restaurant_id == restaurant_id, RestaurantStation.name == st_name).one_or_none()
                if ok:
                    mi.station = st_name
                else:
                    # ignore invalid station but record warning
                    details.append({"row": row, "index": idx, "status": "updated", "warning": "invalid_station"}); updated += 1; continue
            updated += 1
            details.append({"row": row, "index": idx, "status": "updated"})
        except Exception as e:
            details.append({"row": row, "index": idx, "status": "error", "reason": e.__class__.__name__})
    if str(format).lower() == 'csv':
        import csv
        from io import StringIO
        buf = StringIO(); w = csv.writer(buf)
        w.writerow(["index","status","reason","warning","id","name"])
        for d in details:
            r = d.get('row', {})
            w.writerow([d.get('index',''), d.get('status',''), d.get('reason',''), d.get('warning',''), r.get('id',''), r.get('name','')])
        return Response(buf.getvalue(), media_type='text/csv')
    return {"updated": updated, "rows": len(details), "details": details}


@router.get("/restaurants/{restaurant_id}/bulk_stock_template.xlsx")
def bulk_stock_template_xlsx(restaurant_id: str, user: User = Depends(get_current_user)):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX export unavailable")
    from openpyxl import Workbook  # type: ignore
    wb = Workbook(); ws = wb.active; ws.title = "BulkStock"
    ws.append(["id","name","stock_qty","available","visible","oos_until","station"])
    out = BytesIO(); wb.save(out)
    return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/restaurants/{restaurant_id}/bulk_stock.xlsx")
async def bulk_stock_xlsx(restaurant_id: str, request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db), format: str = 'json'):
    _ensure_operator(db, user.id, min_role="manager")
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX import unavailable")
    body = await request.body()
    if not body:
        raise HTTPException(status_code=400, detail="Empty XLSX body")
    from openpyxl import load_workbook  # type: ignore
    import io
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["id","name","stock_qty","available","visible","oos_until","station"]
    for h in ["id","name"]:
        if h not in idx:
            raise HTTPException(status_code=400, detail=f"Missing column: {h}")
    updated = 0; details = []
    row_idx = 2
    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        data = {h: (vals[idx[h]] if h in idx and idx[h] < len(vals) else None) for h in required}
        mid = (str(data.get('id') or '').strip())
        name = (str(data.get('name') or '').strip())
        mi = db.get(MenuItem, mid) if mid else None
        if not mi and name:
            mi = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id, MenuItem.name == name).one_or_none()
        if not mi:
            details.append({"row": data, "index": row_idx, "status": "skipped", "reason": "item_not_found"}); row_idx += 1; continue
        try:
            if data.get('stock_qty') not in (None, ''):
                mi.stock_qty = int(data['stock_qty'])
            if data.get('available') not in (None, ''):
                mi.available = str(data['available']).lower() in ('1','true','yes','y')
            if data.get('visible') not in (None, ''):
                mi.visible = str(data['visible']).lower() in ('1','true','yes','y')
            if data.get('oos_until'):
                try:
                    from datetime import datetime as _dt
                    if isinstance(data['oos_until'], _dt):
                        mi.oos_until = data['oos_until']
                    else:
                        mi.oos_until = _dt.fromisoformat(str(data['oos_until']).replace('Z','+00:00'))
                except Exception:
                    pass
            if data.get('station'):
                st_name = str(data['station']).strip()
                ok = db.query(RestaurantStation).filter(RestaurantStation.restaurant_id == restaurant_id, RestaurantStation.name == st_name).one_or_none()
                if ok:
                    mi.station = st_name
                else:
                    details.append({"row": data, "index": row_idx, "status": "updated", "warning": "invalid_station"}); updated += 1; row_idx += 1; continue
            updated += 1
            details.append({"row": data, "index": row_idx, "status": "updated"})
        except Exception as e:
            details.append({"row": data, "index": row_idx, "status": "error", "reason": e.__class__.__name__})
        row_idx += 1
    if str(format).lower() == 'csv':
        import csv
        from io import StringIO
        buf = StringIO(); w = csv.writer(buf)
        w.writerow(["index","status","reason","warning","id","name"])
        for d in details:
            r = d.get('row', {})
            w.writerow([d.get('index',''), d.get('status',''), d.get('reason',''), d.get('warning',''), r.get('id',''), r.get('name','')])
        return Response(buf.getvalue(), media_type='text/csv')
    return {"updated": updated, "rows": len(details), "details": details}


# --- New: Orders management (platform operator) ---


ALLOWED_TRANSITIONS = {
    "created": {"accepted", "canceled"},
    "accepted": {"preparing", "canceled"},
    "preparing": {"out_for_delivery", "canceled"},
    "out_for_delivery": {"delivered", "canceled"},
}


@router.get("/orders/{order_id}", response_model=OrderOut)
def op_get_order(order_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    o = db.get(Order, order_id)
    if not o:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    return OrderOut(
        id=str(o.id),
        status=o.status,
        restaurant_id=str(o.restaurant_id),
        total_cents=o.total_cents,
        delivery_address=o.delivery_address,
        created_at=o.created_at,
        payment_request_id=o.payment_request_id,
        items=[
            OrderItemOut(
                menu_item_id=str(oi.menu_item_id),
                name=oi.name_snapshot,
                qty=oi.qty,
                price_cents=oi.price_cents_snapshot,
                subtotal_cents=oi.subtotal_cents,
            )
            for oi in o.items
        ],
    )


@router.post("/orders/{order_id}/status")
def op_update_order_status(order_id: str, status_value: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    o = db.get(Order, order_id)
    if not o:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    if o.status in ("delivered", "canceled"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Final state")
    allowed = ALLOWED_TRANSITIONS.get(o.status, set())
    if status_value not in allowed:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid transition")
    prev = o.status
    o.status = status_value
    # Stamp status timeline
    from datetime import datetime as _dt
    now = _dt.utcnow()
    o.last_status_at = now
    if status_value == "accepted":
        o.accepted_at = now
    elif status_value == "preparing":
        o.preparing_at = now
    elif status_value == "out_for_delivery":
        o.out_for_delivery_at = now
    elif status_value == "delivered":
        o.delivered_at = now
    elif status_value == "canceled":
        o.canceled_at = now
    try:
        notify("food.order.status_changed", {"order_id": str(o.id), "status": o.status})
        send_webhooks(db, "food.order.status_changed", {"order_id": str(o.id), "status": o.status})
    except Exception:
        pass
    # If canceled after acceptance, request refund
    if o.status == "canceled" and prev in ("accepted", "preparing", "out_for_delivery") and o.payment_transfer_id:
        try:
            send_webhooks(
                db,
                "refund.requested",
                {"order_id": str(o.id), "transfer_id": o.payment_transfer_id, "amount_cents": int(o.total_cents)},
            )
            o.refund_status = "requested"
            db.flush()
        except Exception:
            pass
    return {"detail": "ok"}


# --- New: Operator members management ---


@router.get("/members")
def list_members(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    mem = _ensure_operator(db, user.id, min_role="admin")
    rows = db.query(OperatorMember).order_by(OperatorMember.created_at.asc()).all()
    out = []
    for m in rows:
        u = db.get(User, m.user_id)
        out.append({
            "id": str(m.id),
            "user_id": str(m.user_id),
            "role": m.role,
            "user_phone": u.phone if u else None,
            "user_name": u.name if u else None,
            "created_at": m.created_at,
        })
    return out


@router.post("/members")
def add_member(phone: str, role: str = "agent", user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    if role not in ("admin", "manager", "agent", "kitchen", "cashier", "support"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid role")
    u = db.query(User).filter(User.phone == phone).one_or_none()
    if u is None:
        u = User(phone=phone)
        db.add(u)
        db.flush()
    existing = db.query(OperatorMember).filter(OperatorMember.user_id == u.id).one_or_none()
    if existing:
        existing.role = role
        return {"id": str(existing.id), "detail": "updated"}
    m = OperatorMember(user_id=u.id, role=role)
    db.add(m)
    db.flush()
    return {"id": str(m.id), "detail": "created"}


@router.patch("/members/{member_id}")
def update_member(member_id: str, role: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    if role not in ("admin", "manager", "agent", "kitchen", "cashier", "support"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid role")
    m = db.get(OperatorMember, member_id)
    if not m:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not found")
    m.role = role
    return {"detail": "ok"}


@router.delete("/members/{member_id}")
def delete_member(member_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    m = db.get(OperatorMember, member_id)
    if not m:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not found")
    db.delete(m)
    return {"detail": "ok"}


# --- New: Advanced reports ---


@router.get("/reports/top_items")
def top_items(days: int = 30, limit: int = 10, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow()
    fr = to - timedelta(days=max(1, min(days, 180)))
    # Aggregate in Python over recent orders; robust to later menu edits via name_snapshot
    from collections import Counter
    cnt: Counter[str] = Counter()
    for o in db.query(Order).filter(Order.created_at >= fr, Order.created_at <= to).all():
        for it in o.items:
            key = it.name_snapshot.strip().lower()
            cnt[key] += it.qty
    top = cnt.most_common(limit)
    return [{"name": name, "qty": qty} for name, qty in top]


@router.get("/reports/cancellation_rate")
def cancellation_rate(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow(); fr = to - timedelta(days=max(1, min(days, 180)))
    total = db.query(func.count(Order.id)).filter(Order.created_at >= fr, Order.created_at <= to).scalar() or 0
    canceled = db.query(func.count(Order.id)).filter(Order.created_at >= fr, Order.created_at <= to, Order.status == "canceled").scalar() or 0
    rate = (float(canceled) / float(total)) if total else 0.0
    return {"from_utc": fr.isoformat()+"Z", "to_utc": to.isoformat()+"Z", "orders_total": int(total), "orders_canceled": int(canceled), "cancellation_rate": rate}


@router.get("/reports/sla")
def sla(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow(); fr = to - timedelta(days=max(1, min(days, 180)))
    rows = db.query(Order).filter(Order.created_at >= fr, Order.created_at <= to).all()
    import statistics as _st
    acc_durs = []
    cook_durs = []
    ship_durs = []
    total_durs = []
    for o in rows:
        if o.accepted_at and o.created_at:
            acc_durs.append((o.accepted_at - o.created_at).total_seconds())
        if o.out_for_delivery_at and o.preparing_at:
            cook_durs.append((o.out_for_delivery_at - o.preparing_at).total_seconds())
        if o.delivered_at and o.out_for_delivery_at:
            ship_durs.append((o.delivered_at - o.out_for_delivery_at).total_seconds())
        if o.delivered_at and o.created_at:
            total_durs.append((o.delivered_at - o.created_at).total_seconds())
    def _stats(a):
        return {"count": len(a), "avg_secs": (sum(a)/len(a) if a else 0), "p50_secs": (_st.median(a) if a else 0), "p90_secs": (_st.quantiles(a, n=10)[8] if len(a) >= 10 else (_st.median(a) if a else 0))}
    return {
        "from_utc": fr.isoformat()+"Z",
        "to_utc": to.isoformat()+"Z",
        "accepted_latency": _stats(acc_durs),
        "preparing_to_out_latency": _stats(cook_durs),
        "out_to_delivered_latency": _stats(ship_durs),
        "created_to_delivered_latency": _stats(total_durs),
    }


@router.get("/reports/peaks")
def peaks(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow(); fr = to - timedelta(days=max(1, min(days, 180)))
    rows = db.query(Order.created_at).filter(Order.created_at >= fr, Order.created_at <= to).all()
    from collections import Counter
    hours = Counter(); dows = Counter()
    for (dt,) in rows:
        if not dt:
            continue
        h = dt.hour; w = dt.weekday()
        hours[h] += 1
        dows[w] += 1
    return {
        "from_utc": fr.isoformat()+"Z",
        "to_utc": to.isoformat()+"Z",
        "by_hour": {str(h): hours.get(h, 0) for h in range(24)},
        "by_weekday": {str(w): dows.get(w, 0) for w in range(7)},
    }


@router.get("/reports/by_restaurant")
def by_restaurant(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow()
    fr = to - timedelta(days=max(1, min(days, 180)))
    rows = (
        db.query(
            Restaurant.id.label("restaurant_id"),
            Restaurant.name.label("restaurant_name"),
            func.count(Order.id).label("orders"),
            func.sum(Order.total_cents).label("revenue_cents"),
        )
        .join(Order, Order.restaurant_id == Restaurant.id)
        .filter(Order.created_at >= fr, Order.created_at <= to)
        .group_by(Restaurant.id, Restaurant.name)
        .order_by(func.sum(Order.total_cents).desc())
        .all()
    )
    return [
        {
            "restaurant_id": str(r.restaurant_id),
            "restaurant_name": r.restaurant_name,
            "orders": int(r.orders or 0),
            "revenue_cents": int(r.revenue_cents or 0),
        }
        for r in rows
    ]


@router.get("/reports/by_restaurant.xlsx")
def by_restaurant_xlsx(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX export unavailable")
    rows = by_restaurant(days, user, db)
    from openpyxl import Workbook  # type: ignore
    wb = Workbook(); ws = wb.active; ws.title = "By Restaurant"
    ws.append(["restaurant_id","restaurant_name","orders","revenue_cents"])
    for r in rows:
        ws.append([r["restaurant_id"], r["restaurant_name"], r["orders"], r["revenue_cents"]])
    out = BytesIO(); wb.save(out)
    return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/reports/payout")
def payout_report(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow(); fr = to - timedelta(days=max(1, min(days, 90)))
    rows = (
        db.query(
            Restaurant.id.label("restaurant_id"),
            Restaurant.name.label("restaurant_name"),
            func.sum(Order.total_cents).label("gross"),
        ).join(Order, Order.restaurant_id == Restaurant.id)
        .filter(Order.created_at >= fr, Order.created_at <= to, Order.status == "delivered")
        .group_by(Restaurant.id, Restaurant.name)
        .all()
    )
    fee_bps = int(getattr(settings, "PLATFORM_FEE_BPS", 0))
    tax_bps = int(getattr(settings, "TAX_RATE_BPS", 0))
    out_rows = []
    total_gross = total_fee = total_tax = total_net = 0
    for r in rows:
        gross = int(r.gross or 0)
        fee = gross * fee_bps // 10000
        tax = gross * tax_bps // 10000
        net = gross - fee - tax
        total_gross += gross; total_fee += fee; total_tax += tax; total_net += net
        out_rows.append({
            "restaurant_id": str(r.restaurant_id),
            "restaurant_name": r.restaurant_name,
            "gross_cents": gross,
            "fee_cents": fee,
            "tax_cents": tax,
            "net_payout_cents": net,
        })
    return {
        "from_utc": fr.isoformat()+"Z",
        "to_utc": to.isoformat()+"Z",
        "fee_bps": fee_bps,
        "tax_bps": tax_bps,
        "totals": {
            "gross_cents": total_gross,
            "fee_cents": total_fee,
            "tax_cents": total_tax,
            "net_payout_cents": total_net,
        },
        "rows": out_rows,
    }


@router.get("/reports/payout.xlsx")
def payout_report_xlsx(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="XLSX export unavailable")
    data = payout_report(days, user, db)
    from openpyxl import Workbook  # type: ignore
    wb = Workbook(); ws = wb.active; ws.title = "Payout"
    ws.append(["from_utc","to_utc","fee_bps","tax_bps"]) ; ws.append([data["from_utc"], data["to_utc"], data["fee_bps"], data["tax_bps"]])
    ws2 = wb.create_sheet("Rows"); ws2.append(["restaurant_id","restaurant_name","gross_cents","fee_cents","tax_cents","net_payout_cents"])
    for r in data["rows"]:
        ws2.append([r["restaurant_id"], r["restaurant_name"], r["gross_cents"], r["fee_cents"], r["tax_cents"], r["net_payout_cents"]])
    out = BytesIO(); wb.save(out)
    return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/reports/payout.pdf")
def payout_report_pdf(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    data = payout_report(days, user, db)
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF unavailable: {e}")
    buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4); width, height = A4
    y = height - 72
    c.setFont("Helvetica-Bold", 16); c.drawString(72, y, "Payout Report"); y -= 20
    c.setFont("Helvetica", 11); c.drawString(72, y, f"Range: {data['from_utc']} .. {data['to_utc']}"); y -= 16
    c.drawString(72, y, f"Fee BPS: {data['fee_bps']}  Tax BPS: {data['tax_bps']}"); y -= 20
    c.setFont("Helvetica-Bold", 12); c.drawString(72, y, "restaurant_name            gross    fee    tax    net"); y -= 14
    c.setFont("Helvetica", 11)
    for r in data.get("rows", [])[:40]:
        rn = (r['restaurant_name'] or '')[:24].ljust(24)
        c.drawString(72, y, f"{rn}  {r['gross_cents']:>7}  {r['fee_cents']:>6}  {r['tax_cents']:>6}  {r['net_payout_cents']:>7}"); y -= 14
        if y < 100:
            c.showPage(); y = height - 72; c.setFont("Helvetica", 11)
    y -= 10; c.setFont("Helvetica-Bold", 12)
    t=data.get('totals',{})
    c.drawString(72, y, f"Totals: gross={t.get('gross_cents',0)} fee={t.get('fee_cents',0)} tax={t.get('tax_cents',0)} net={t.get('net_payout_cents',0)}")
    c.showPage(); c.save()
    return Response(buf.getvalue(), media_type="application/pdf")


@router.get("/reports/by_restaurant.csv")
def by_restaurant_csv(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = by_restaurant(days, user, db)
    buf = StringIO(); w = csv.writer(buf)
    w.writerow(["restaurant_id","restaurant_name","orders","revenue_cents"])
    for r in rows:
        w.writerow([r["restaurant_id"], r["restaurant_name"], r["orders"], r["revenue_cents"]])
    return Response(buf.getvalue(), media_type="text/csv")


@router.get("/reports/by_city")
def by_city(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="agent")
    to = datetime.utcnow()
    fr = to - timedelta(days=max(1, min(days, 180)))
    rows = (
        db.query(
            Restaurant.city.label("city"),
            func.count(Order.id).label("orders"),
            func.sum(Order.total_cents).label("revenue_cents"),
        )
        .join(Order, Order.restaurant_id == Restaurant.id)
        .filter(Order.created_at >= fr, Order.created_at <= to)
        .group_by(Restaurant.city)
        .order_by(func.sum(Order.total_cents).desc())
        .all()
    )
    return [
        {
            "city": r.city,
            "orders": int(r.orders or 0),
            "revenue_cents": int(r.revenue_cents or 0),
        }
        for r in rows
    ]


@router.get("/reports/by_city.csv")
def by_city_csv(days: int = 30, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = by_city(days, user, db)
    buf = StringIO(); w = csv.writer(buf)
    w.writerow(["city","orders","revenue_cents"])
    for r in rows:
        w.writerow([r["city"], r["orders"], r["revenue_cents"]])
    return Response(buf.getvalue(), media_type="text/csv")


@router.post("/webhooks/retry")
def webhooks_retry(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="admin")
    try:
        process_pending_once()
        return {"detail": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- KDS (Operator-level; kitchen role) ---


@router.get("/kds/orders")
def kds_orders(status: str = "accepted", station: str | None = None, limit: int = 100, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="kitchen")
    st = [status] if status else ["accepted", "preparing"]
    rows = db.query(Order).filter(Order.status.in_(st)).order_by(Order.created_at.asc()).limit(max(1, min(limit, 200))).all()
    out = []
    for o in rows:
        items = [{"name": it.name_snapshot, "qty": it.qty, "packed": it.packed, "station": getattr(it, 'station_snapshot', None)} for it in o.items]
        if station:
            items = [it for it in items if (it.get('station') or '') == station]
        if not items:
            continue
        out.append({"id": str(o.id), "status": o.status, "created_at": o.created_at, "items": items})
    return out


@router.post("/kds/orders/{order_id}/bump")
def kds_bump(order_id: str, to_status: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="kitchen")
    o = db.get(Order, order_id)
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    o.kds_bumped = True
    if to_status in ("preparing", "out_for_delivery", "delivered"):
        allowed = ALLOWED_TRANSITIONS.get(o.status, set())
        if to_status not in allowed:
            raise HTTPException(status_code=400, detail="Invalid transition")
        o.status = to_status
    # Best-effort ESC/POS print
    try:
        lines = [f"Order {o.id}", f"Status: {o.status}"] + [f"- {it.qty} x {it.name_snapshot}" for it in o.items]
        print_ticket(lines)
    except Exception:
        pass
    return {"detail": "ok"}


@router.get("/kds/orders/{order_id}/ticket.pdf")
def kds_ticket_pdf(order_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ensure_operator(db, user.id, min_role="kitchen")
    o = db.get(Order, order_id)
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A6
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF unavailable: {e}")
    buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A6); width, height = A6
    y = height - 36
    c.setFont("Helvetica-Bold", 12); c.drawString(20, y, f"KDS Ticket"); y -= 14
    c.setFont("Helvetica", 10); c.drawString(20, y, f"Order: {o.id}"); y -= 12
    c.drawString(20, y, f"Status: {o.status}"); y -= 12
    c.drawString(20, y, f"Created: {o.created_at}"); y -= 16
    c.setFont("Helvetica-Bold", 11); c.drawString(20, y, "Items:"); y -= 12
    c.setFont("Helvetica", 10)
    for it in o.items:
        line = f"{it.qty} x {it.name_snapshot}"
        if getattr(it, 'station_snapshot', None):
            line += f" [{it.station_snapshot}]"
        c.drawString(22, y, line); y -= 12
        if y < 40:
            c.showPage(); y = height - 36; c.setFont("Helvetica", 10)
    c.showPage(); c.save()
    return Response(buf.getvalue(), media_type="application/pdf")
